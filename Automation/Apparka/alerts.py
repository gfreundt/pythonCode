import scrapers
import json, csv
import openpyxl
from copy import deepcopy as copy
import time
from gft_utils import ChromeUtils, GoogleUtils
import threading
from datetime import datetime as dt, timedelta as td
import logging
from jinja2 import Environment, FileSystemLoader
from pprint import pprint
import uuid


def load_members():
    with open("members2.json", mode="r", encoding="utf-8") as file:
        return json.load(file)


def save_members(members):
    with open("members2.json", mode="w+", encoding="utf-8") as file:
        json.dump(members, file, indent=4)
    LOG.info("Database updated sccuesfully")


def load_activity():
    with open("activity.csv", mode="r") as file:
        return list(csv.reader(file))


def load_raw_members():
    # download latest form responses
    G = GoogleUtils()
    _folder_id = "1Az6eM7Fr9MUqNhj-JtvOa6WOhYBg5Qbs"
    _file = [i for i in G.get_drive_files(_folder_id) if "members" in i["title"]][0]
    G.download_from_drive(_file, ".\data\members.xlsx")

    # open latest form responses
    wb = openpyxl.load_workbook(".\data\members.xlsx")
    sheet = wb["Form Responses 2"]
    raw_data = [
        {
            sheet.cell(row=1, column=j).value: sheet.cell(row=i + 1, column=j).value
            for j in range(2, sheet.max_column + 1)
        }
        for i in range(1, sheet.max_row)
    ]

    # clean data and join placas in one list
    for r, row in enumerate(copy(raw_data)):
        placas = []
        for col in row:
            if type(row[col]) == float:
                raw_data[r][col] = str(int(row[col]))
            if type(raw_data[r][col]) == str and "@" in raw_data[r][col]:
                raw_data[r][col] = raw_data[r][col].lower()
            if "Nombre" in col:
                raw_data[r][col] = raw_data[r][col].title()
            if "Correo" in col:
                raw_data[r]["Correo"] = raw_data[r].pop(col)
            if "Placa" in col:
                if raw_data[r][col]:
                    placas.append(
                        raw_data[r][col].replace("-", "").replace(" ", "").upper()
                    )
                del raw_data[r][col]
        raw_data[r].update(
            {
                "Placas": placas,
            }
        )
    return raw_data


def add_new_members(all_members, existing_members):
    correlativo = len(existing_members)
    existing_members_docs = [
        i["Datos"]["Número de Documento"] for i in existing_members
    ]
    for all_member in all_members:
        if str(all_member["Número de Documento"]) not in existing_members_docs:
            _new_record = {
                "Correlativo": correlativo,
                "Codigo": "SAP-" + str(uuid.uuid4())[-6:].upper(),
                "Datos": all_member,
                "Resultados": {
                    "Satimp": [],
                    "Satimp_Actualizado": "01/01/1999",
                    "Brevete": {},
                    "Brevete_Actualizado": "01/01/1999",
                    "Revtec": [],
                    "Revtec_Actualizado": "01/01/1999",
                    "Sutran": [],
                    "Sutran_Actualizado": "01/01/1999",
                },
                "Envios": {
                    "Bienvenida": {},
                    "Resumen": [],
                    "Alertas": {
                        "Brevete": [],
                        "Satimp": [],
                        "Revtec": [],
                        "Sutran": [],
                    },
                },
            }

            correlativo += 1
            existing_members.append(_new_record)
            LOG.info(
                f"Appended new record: {_new_record['Correlativo']} - {_new_record['Datos']['Nombre y Apellido']}"
            )

    return existing_members


def get_records_to_process(members):
    # TODO: beyond welcome
    docs_to_process, placas_to_process = [], []
    for member in members:
        if not member["Envios"]["Bienvenida"]:
            docs_to_process.append(
                (
                    member["Correlativo"],
                    -1,
                    member["Datos"]["Documento Tipo"],
                    member["Datos"]["Número de Documento"],
                    "",
                )
            )
            for k, placa in enumerate(member["Datos"]["Placas"]):
                placas_to_process.append((member["Correlativo"], k, "", "", placa))

    return docs_to_process, placas_to_process


def gather_soat(members, placas_to_process):
    URL = "https://www.apeseg.org.pe/consultas-soat/"
    SOAT = scrapers.Soat()
    SOAT.WEBD = ChromeUtils().init_driver(headless=True, verbose=False, maximized=True)
    SOAT.WEBD.get(URL)
    time.sleep(2)
    responses = []
    for data in placas_to_process:
        captcha_txt = ""
        img = SOAT.browser(placa=data[4], captcha_txt=captcha_txt)
        img.show()
        captcha_txt = input("Captcha: ")
        responses.append(
            (data[0], data[1], SOAT.browser(placa=data[4], captcha_txt=captcha_txt))
        )

    pprint(responses)
    new_response = [{"SOAT": [[], [], []]} for _ in range(len(members))]
    for rec, pos, data in responses:
        new_response[rec]["SOAT"][pos] = data

    # delete dictionary keys that have no data to update
    return [
        {k: v for k, v in j.items() if v and v != [[], [], []]} for j in new_response
    ]


def gather(members, docs_to_process, placas_to_process):
    URLS = [
        "https://www.sat.gob.pe/WebSitev8/IncioOV2.aspx",
        "https://licencias.mtc.gob.pe/#/index",
        "https://portal.mtc.gob.pe/reportedgtt/form/frmconsultaplacaitv.aspx",
        "https://www.sutran.gob.pe/consultas/record-de-infracciones/record-de-infracciones/",
    ]
    scraper = [
        (scrapers.Satimp(None), docs_to_process),
        (scrapers.Brevete(None, URLS[1]), docs_to_process),
        (scrapers.Revtec(None), placas_to_process),
        (scrapers.Sutran(None), placas_to_process),
    ]
    threads = []
    for index, (url, scrap) in enumerate(zip(URLS, scraper)):
        _t = threading.Thread(target=full_update, args=(url, scrap[0], scrap[1], index))
        threads.append(_t)
        _t.start()

    for thread in threads:
        thread.join()

    # join responses in member list order
    new_response = [
        {"Satimp": {}, "Brevete": {}, "Revtec": [[], [], []], "Sutran": [[], [], []]}
        for _ in range(len(members))
    ]
    for response, header in zip(responses, HEADERS):
        for resp in response:
            rec, pos, data = resp
            if pos == -1:
                new_response[rec].update({header: data})
            else:
                new_response[rec][header][pos] = data

    # delete dictionary keys that have no data to update
    # TODO: Sutran?
    return [
        {k: v for k, v in j.items() if v and v != [[], [], []]} for j in new_response
    ]


def full_update(URL, scraper, data_to_process, index):
    # define Chromedriver and open url for first time
    scraper.WEBD = ChromeUtils().init_driver(
        headless=True, verbose=False, maximized=True
    )
    scraper.WEBD.get(URL)
    time.sleep(3)

    # iterate on all records that require updating
    for k, data in enumerate(data_to_process):
        print(f"{index}: {(k+1)/len(data_to_process)*100:.1f}%")
        # get scraper data
        try:
            attempts, new_record = 0, ""
            while not new_record and attempts < 5:
                new_record, _ = scraper.browser(
                    doc_tipo=data[2], doc_num=data[3], placa=data[4]
                )
                responses[index].append((data[0], data[1], new_record))
                attempts += 1
        except KeyboardInterrupt:
            quit()
        except:
            LOG.warning(f"No proceso {data}")
            responses[index].append((data[0], data[1], ""))


def update_member_data(members, new_responses):
    for k, response in enumerate(new_responses):
        members[k]["Resultados"].update(response)
        for header in HEADERS:
            members[k]["Resultados"][f"{header}_Actualizado"] = dt.now().strftime(
                "%d/%m/%Y"
            )
    return members


def get_alert_lists(members):

    BREVETE_CHECKPOINTS = (-30, -15, -5, 0)
    REVTEC_CHECKPOINTS = (-10, -5, 0)
    CURRENTQ = (dt.now().month - 1) // 3 + 1
    welcome_list, regular_list = [], []
    warning_list = [[] for i in range(4)]  # order: brevete, satimp, revtec, sutran

    for j, member in enumerate(members):
        # add to Welcome List (do not consider any other lists)
        if not member["Envios"]["Bienvenida"]:
            welcome_list.append(j)
            # continue

    return welcome_list, regular_list, warning_list
    """

        # add to Regular List
        if not member["Envios"]["Resumen"] or dt.now() - dt.strptime(
            member["Envios"]["Resumen"], "%d/%m/%Y"
        ) >= td(days=30):
            regular_list.append(j)

        # add to Warning List - Brevete
        if (
            _m
            and (dt.now() - dt.strptime(_m["fecha_hasta"], "%d/%m/%Y")).days
            in BREVETE_CHECKPOINTS
        ):
            warning_list[0].append(j, 0, 0)

        # add to Warning List - Satimp
        for m, codigo in enumerate(member["Resultados"]["Satimp"]):
            for n, deuda in enumerate(codigo["deudas"]):
                if int(deuda["ano"]) < dt.now().year or (
                    int(deuda["ano"]) == dt.now().year
                    and int(deuda["periodo"]) <= CURRENTQ
                ):
                    warning_list[1].append((j, m, n))

        for k, (revtec, sutran) in enumerate(
            zip(member["Resultados"]["Revtec"], member["Resultados"]["Sutran"])
        ):
            # add to Warning List - Revision Tecnica
            if (
                revtec
                and (dt.now() - dt.strptime(revtec[0]["fecha_hasta"], "%d/%m/%Y")).days
                in REVTEC_CHECKPOINTS
            ):
                warning_list[2].append((j, k, 0))

            # add to Warning List - Sutran
            if sutran:
                warning_list[3].append((j, k, 0))

    return welcome_list, regular_list, warning_list
    """


def send_alerts(members, welcome_list, regular_list, warning_list):
    CURRENTQ = (dt.now().month - 1) // 3 + 1
    messages = []

    # Welcome Message
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("bienvenida.html")

    for mem, selected_member in enumerate(welcome_list):
        member = members[selected_member]
        _combina_deudas_sat = [
            i["deudas"] for i in member["Resultados"]["Satimp"] if i["deudas"]
        ]

        if _combina_deudas_sat:
            _combina_deudas_sat = [
                (int(i["ano"]), int(i["periodo"])) for i in _combina_deudas_sat[0] if i
            ]

        _alertas = [
            (
                "Licencia de Conducir vencida o vence en menos de 30 días."
                if member["Resultados"]["Brevete"]
                and dt.strptime(
                    member["Resultados"]["Brevete"]["fecha_hasta"], "%d/%m/%Y"
                )
                - dt.now()
                <= td(days=30)
                else ""
            ),
            (
                "Al menos una Revision Tecnica vencida o vence en menos de 15 días."
                if any(
                    [
                        member["Resultados"]["Revtec"]
                        and dt.strptime(i[0]["fecha_hasta"], "%d/%m/%Y") - dt.now()
                        <= td(days=15)
                        for i in member["Resultados"]["Revtec"]
                        if i
                    ]
                )
                else ""
            ),
            (
                "Impuesto Vehicular SAT vencido o en periodo de pago."
                if any(
                    [
                        i[0] < dt.now().year
                        or (i[0] == dt.now().year and i[1] <= CURRENTQ)
                        for i in _combina_deudas_sat
                    ]
                )
                else ""
            ),
            (
                "Multa impaga en SUTRAN."
                if any(i for i in member["Resultados"]["Sutran"])
                else ""
            ),
        ]

        _alertas = [i for i in _alertas if i]
        _info = {"alertas": _alertas if _alertas else ["Ninguna"]}
        email_id = f"{member['Codigo']}|{str(uuid.uuid4())[-12:]}"
        _info.update(
            {
                "nombre_usuario": member["Datos"]["Nombre y Apellido"],
                "codigo_correo": email_id,
            }
        )
        _revtecs = []
        for _m in member["Resultados"]["Revtec"]:
            if _m:
                _revtecs.append(
                    {
                        "certificadora": _m[0]["certificadora"].split("-")[-1][:35],
                        "placa": _m[0]["placa"],
                        "certificado": _m[0]["certificado"],
                        "fecha_desde": _m[0]["fecha_desde"],
                        "fecha_hasta": _m[0]["fecha_hasta"],
                        "resultado": _m[0]["resultado"],
                        "vigencia": _m[0]["vigencia"],
                    }
                )
        _info.update({"revtecs": _revtecs})

        _m = member["Resultados"]["Brevete"]
        if _m:
            _info.update(
                {
                    "brevete": {
                        "numero": _m["numero"],
                        "clase": _m["clase"],
                        "formato": _m["tipo"],
                        "fecha_desde": _m["fecha_expedicion"],
                        "fecha_hasta": _m["fecha_hasta"],
                        "restricciones": _m["restricciones"],
                        "local": _m["centro"],
                    }
                }
            )
        else:
            _info.update({"brevete": {}})

        _sutran = []
        for k, _m in enumerate(member["Resultados"]["Sutran"]):
            if _m:
                _sutran.append(
                    {
                        "placa": member["Datos"]["Placas"][k],
                        "documento": _m["documento"] if _m else [],
                        "tipo": _m["tipo"] if _m else [],
                        "fecha_documento": _m["fecha_documento"] if _m else [],
                        "infraccion": (
                            f"{_m['codigo_infraccion']} - {_m['clasificacion']}"
                            if _m
                            else []
                        ),
                    }
                )
        _info.update({"sutrans": _sutran})

        _m = member["Resultados"]["Satimp"]
        if _m:
            _info.update(
                {
                    "satimp": {
                        "codigo": _m[0]["codigo"],
                        "deudas": _m[0]["deudas"],
                    }
                }
            )
        else:
            _info.update({"satimp": {}})

        content = template.render(_info)
        messages.append(
            {
                "to": member["Datos"]["Correo"],
                "cc": "gabfre@gmail.com",
                "subject": "Bienvenido al Servicio de Alertas Perú",
                "body": content,
                "attachments": [],
            }
        )
        members[selected_member]["Envios"]["Bienvenida"].update(
            {"fecha": dt.now().strftime("%d/%m/%Y"), "hash": email_id}
        )

        LOG.info(
            f'Email included to {member["Datos"]["Nombre"]} ({member["Datos"]["Documento Tipo"]}:{member["Datos"]["Número de Documento"]})'
        )

    with open("messages.txt", "w") as file:
        for i in messages:
            file.write(i["body"])
            file.write("=" * 50)

    # GoogleUtils().send_gmail(fr="servicioalertaperu@gmail.com", messages=messages)
    LOG.info("Emails sent succesfully")

    return members


def start_logger():
    _filename = "alerts_log.txt"
    logging.basicConfig(
        filename=_filename,
        filemode="w",
        format="%(asctime)s | %(levelname)s | %(message)s",
    )
    _log = logging.getLogger(__name__)
    _log.setLevel(logging.INFO)
    return _log


def main():
    # # # Step 1: Download raw list of all members from form and add new ones with default data
    # members = add_new_members(
    #     all_members=load_raw_members(), existing_members=load_members()
    # )
    members = load_members()
    # # # Step 2A: Scraping (get list of records to process)
    docs_to_process, placas_to_process = get_records_to_process(members)
    # # # Step 2B: Scraping (activate scrapers)
    # new_responses = gather(members, docs_to_process, placas_to_process)
    new_responses = gather_soat(members, placas_to_process)
    # # Step 3: Update data (new information from scrapers into database)
    members = update_member_data(members, new_responses)
    # Step 4A: Alerting (get list of records to process)
    # welcome_list, regular_list, warning_list = get_alert_lists(members)
    # Step 4B: Alerting (send alerts by list)
    # members = send_alerts(members, welcome_list, regular_list, warning_list)
    # Step 5: Update file with new scraped information and data from allsent messages
    save_members(members)


# G = "1Az6eM7Fr9MUqNhj-JtvOa6WOhYBg5Qbs"
# GoogleUtils().upload_to_drive(
#     local_path="members.xlsx", gdrive_filename="members", gdrive_path_id=G
# )
# quit()
HEADERS = ["Satimp", "Brevete", "Revtec", "Sutran"]
responses = [[] for _ in range(4)]
LOG = start_logger()
main()
