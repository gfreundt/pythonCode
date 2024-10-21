import uuid
from copy import deepcopy as copy
import openpyxl
from gft_utils import GoogleUtils


def add(db_cursor, monitor):

    # log start of process
    monitor.log("Checking for New Members.")

    # download online form and add to local database
    form_members = load_form_members()

    # get DocNum from existing members as index to avoid duplicates
    db_cursor.execute("SELECT DocNum FROM members")
    existing_members_docs = [i[0] for i in db_cursor.fetchall()]

    # iterate on all members from online form and only add new ones (new DocNum)
    counter = 0
    for member in form_members:

        # add basic information to members table
        if str(member["Número de Documento"]) not in existing_members_docs:
            counter += 1
            cmd = format_query(
                "members",
                [
                    "NombreCompleto",
                    "DocTipo",
                    "DocNum",
                    "Celular",
                    "WhatsApp",
                    "Correo",
                    "CodMember",
                    "Unsubscribe",
                ],
            )
            values = (
                list(member.values())[:6]
                + ["SAP-" + str(uuid.uuid4())[-6:].upper()]
                + [0]
            )
            db_cursor.execute(cmd, values)

            # get autoindex number to link to placas table
            cmd = f"SELECT * FROM members WHERE DocNum = '{values[2]}'"
            db_cursor.execute(cmd)
            idmember = db_cursor.fetchone()[0]
            monitor.log(
                f"Appended new record: {idmember} - {member['Número de Documento']}"
            )

            # add placas to placas table
            for placa in member["Placas"]:
                cmd = format_query("placas", ["IdMember_FK", "Placa"])
                values = [idmember] + [placa]
                db_cursor.execute(cmd, values)
                monitor.log(f"Added placa: {placa} for MemberID {idmember}")

    # add counter to monitor display
    monitor.log(f"Total Records: {str(counter)}", type=1)


def load_form_members():
    """Download online form responses, select new ones only, clean and structure data."""
    google = GoogleUtils()
    # set Google Drive folder, download entire spreadsheet
    _folder_id = "1Az6eM7Fr9MUqNhj-JtvOa6WOhYBg5Qbs"
    _file = [i for i in google.get_drive_files(_folder_id) if "members" in i["title"]][
        0
    ]
    google.download_from_drive(_file, "..\data\members.xlsx")
    # open spreadsheet, select correct tab and load into list of dictionaries
    wb = openpyxl.load_workbook("..\data\members.xlsx")
    sheet = wb["Form Responses 2"]
    raw_data = [
        {
            sheet.cell(row=1, column=j).value: sheet.cell(row=i + 1, column=j).value
            for j in range(2, sheet.max_column + 1)
        }
        for i in range(1, sheet.max_row)
    ]

    # clean data and shape data correctly (join placas in single list)
    _wa = list(raw_data[0].keys())[8]
    for r, row in enumerate(copy(raw_data)):
        placas = []
        for col in row:
            if type(row[col]) == float:
                raw_data[r][col] = str(int(row[col]))
            if type(raw_data[r][col]) == str and "@" in raw_data[r][col]:
                raw_data[r][col] = raw_data[r][col].lower()
            if "Nombre" in col:
                raw_data[r][col] = raw_data[r][col].title()
            if "Correo" in col:
                raw_data[r]["Correo"] = raw_data[r].pop(col)
            if "WhatsApp" in col:
                raw_data[r][_wa] = 1 if raw_data[r][_wa] == "Sí" else 0
            if "Placa" in col:
                if raw_data[r][col]:
                    placas.append(
                        raw_data[r][col].replace("-", "").replace(" ", "").upper()
                    )
                del raw_data[r][col]
        raw_data[r].update(
            {
                "Placas": placas,
            }
        )
    return raw_data


def format_query(table, fields):
    """Create generic SQL statement to insert new records in any table in database."""

    _qmarks = ",".join(["?" for _ in range(len(fields))])
    # return example: REPLACE INTO members (IdMember, Email) VALUES (?,?)
    return f"REPLACE INTO {table} ({','.join(fields)}) VALUES ({_qmarks})"
