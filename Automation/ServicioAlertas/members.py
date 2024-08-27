import os
import openpyxl
import sqlite3
from copy import deepcopy as copy
import uuid
from gft_utils import GoogleUtils
from pprint import pprint


class Members:
    """Used to download new members and add them to local database, manage unsubscribe
    requests and create 30-day list"""

    def __init__(self, LOG, MONITOR):
        self.LOG = LOG
        self.MONITOR = MONITOR
        self.GOOGLE = GoogleUtils()
        self.GMAIL_ACCOUNT = "servicioalertaperu@gmail.com"

        # connect to database, enable threading
        SQLDATABASE = os.path.join(os.getcwd(), "data", "members.sqlite")
        self.conn = sqlite3.connect(SQLDATABASE, check_same_thread=False)
        self.cursor = self.conn.cursor()

    def add_new_members(self):
        # log start of process
        self.LOG.info("Checking for New Members.")

        # download online form and add to local database
        form_members = self.load_form_members()

        # get DocNum from existing members as index to avoid duplicates
        self.cursor.execute("SELECT DocNum FROM members")
        existing_members_docs = [i[0] for i in self.cursor.fetchall()]

        # iterate on all members from online form and only add new ones (new DocNum)
        counter = 0
        for member in form_members:
            # add basic information to members table
            if str(member["Número de Documento"]) not in existing_members_docs:
                counter += 1
                cmd = self.sql(
                    "members",
                    [
                        "NombreCompleto",
                        "DocTipo",
                        "DocNum",
                        "Celular",
                        "WhatsApp",
                        "Correo",
                        "CodMember",
                        "Unsubscribe",
                    ],
                )
                values = (
                    list(member.values())[:6]
                    + ["SAP-" + str(uuid.uuid4())[-6:].upper()]
                    + [0]
                )
                self.cursor.execute(cmd, values)

                # get autoindex number to link to placas table
                cmd = f"SELECT * FROM members WHERE DocNum = '{values[2]}'"
                self.cursor.execute(cmd)
                idmember = self.cursor.fetchone()[0]
                self.LOG.info(
                    f"Appended new record: {idmember} - {member['Número de Documento']}"
                )

                # add placas to placas table
                for placa in member["Placas"]:
                    cmd = self.sql("placas", ["IdMember_FK", "Placa"])
                    values = [idmember] + [placa]
                    self.cursor.execute(cmd, values)
                    self.LOG.info(f"Added placa: {placa} for MemberID {idmember}")

            self.conn.commit()

        self.MONITOR.add_widget(str(counter), type=2)

    def sub_unsub(self):
        """Check inbox for unsubscribe/resubscribe requests and switch table in database."""

        # log start of process
        self.LOG.info("Checking for Sub/Unsub requests.")

        _inbox = self.GOOGLE.read_gmail(
            fr=self.GMAIL_ACCOUNT, only_unread=True, mark_read=False
        )
        # create lists of email addresses of requests to unsubscribe or resuscribe from unread emails
        unsub_emails = []
        sub_emails = []
        for email in _inbox:
            for m in email.messages:
                # register sender to use as key to change flag
                sender = m.sender.split("<")[1][:-1] if "<" in m.sender else ""
                # load subject and body text
                sub = m.subject.split(f"<{self.GMAIL_ACCOUNT}>")[0] if m.subject else ""
                body = m.body.split(f"<{self.GMAIL_ACCOUNT}>")[0] if m.body else ""
                # check if request in subject or body, not case-sensitive
                if "ELIMINAR" in sub.upper() or "ELIMINAR" in body.upper():
                    unsub_emails.append(sender)
                    email.markAsRead()
                elif "INGRESAR" in sub.upper() or "INGRESAR" in body.upper():
                    sub_emails.append(sender)
                    email.markAsRead()

        # move from active members to past members
        for unsub in unsub_emails:
            cmd = f"INSERT INTO pastMembers SELECT * FROM members WHERE Correo='{unsub}'; DELETE FROM members WHERE Correo='{unsub}'"
            self.cursor.executescript(cmd)
            self.LOG.info(
                f"Removed members with email {unsub} from active member list."
            )
        # move from past members to active members
        for sub in sub_emails:
            cmd = f"INSERT INTO members SELECT * FROM pastMembers WHERE Correo='{unsub}'; DELETE FROM pastMembers WHERE Correo='{unsub}'"
            self.cursor.executescript(cmd)
            self.LOG.info(
                f"Reinstated members with email {unsub} into active member list."
            )
        self.conn.commit()

        # TODO: send confirmation email

    def create_30day_list(self):
        """Creates table with all members with SOAT, REVTEC, SUTRAN, SATMUL, BREVETE or SATIMP
        expired or expiring within 30 days."""

        _cmd = f""" DROP TABLE IF EXISTS _expira30dias;
                    CREATE TABLE _expira30dias (IdMember, Placa, FechaHasta, TipoAlerta, Vencido);

                    INSERT INTO _expira30dias (IdMember,  Placa, FechaHasta, TipoAlerta)
                    SELECT IdMember, Placa, FechaHasta, TipoAlerta FROM members
                    JOIN (
                        SELECT * FROM placas 
                        JOIN (
                        SELECT idplaca_FK, FechaHasta, "SOAT" AS TipoAlerta FROM soats WHERE DATE('now', 'localtime', '30 days') >= FechaHasta
                        UNION
                        SELECT idplaca_FK, FechaHasta, "REVTEC" FROM revtecs WHERE DATE('now', 'localtime', '30 days') >= FechaHasta
                        UNION 
                        SELECT idplaca_FK, "", "SUTRAN" FROM sutrans
                        UNION 
                        SELECT idplaca_FK, "", "SATMUL" FROM satmuls)
                        ON idplaca = IdPlaca_FK)
                    ON IdMember = IdMember_FK;

                    INSERT INTO _expira30dias (IdMember, FechaHasta, TipoAlerta)
                    SELECT IdMember, FechaHasta, TipoAlerta from members 
                        JOIN (
                            SELECT IdMember_FK, FechaHasta, "BREVETE" AS TipoAlerta FROM brevetes WHERE DATE('now', 'localtime', '30 days') >= FechaHasta OR DATE('now', 'localtime', '30 days')= FechaHasta OR DATE('now', 'localtime', '0 days')= FechaHasta
						UNION
							SELECT IdMember_FK, FechaHasta, "SATIMP" AS TipoAlerta FROM satimpDeudas 
							JOIN
							(SELECT * FROM satimpCodigos)
							ON IdCodigo_FK = IdCodigo
							WHERE DATE('now', 'localtime', '30 days') >= FechaHasta
                        UNION
                            SELECT IdMember_FK, "", "MTCPAPELETA" FROM mtcPapeletas)
                    ON IdMember = IdMember_FK;
                    
                    UPDATE _expira30dias SET Vencido = 0;
                    UPDATE _expira30dias SET Vencido = 1 WHERE FechaHasta < DATE('now', 'localtime');
                    """
        self.cursor.executescript(_cmd)
        # assign list to instance variable
        self.cursor.execute("SELECT * FROM _expira30dias")
        self.day30_list = self.cursor.fetchall()
        self.LOG.info("Updated 30-day list.")

        _cmd = """  drop table if exists _newSunarpRequired;
                    CREATE TABLE _newSunarpRequired (IdMember_FK);
                    INSERT INTO _newSunarpRequired (IdMember_FK)
                    select IdMember from members where IdMember not in (
                    select DISTINCT IdMember_FK from mensajes 
                    JOIN
                    (select IdMensaje_FK from mensajeContenidos where IdTipoMensaje_FK = 16)
                    on
                    IdMensaje = IdMensaje_FK
                    where DATE('now','localtime','-1 year')<DATE(Fecha))"""
        self.cursor.executescript(_cmd)
        self.LOG.info("Updated New SUNARP required.")

    def load_form_members(self):
        """Download online form responses, select new ones only, clean and structure data"""
        # set Google Drive folder, download entire spreadsheet
        _folder_id = "1Az6eM7Fr9MUqNhj-JtvOa6WOhYBg5Qbs"
        _file = [
            i
            for i in self.GOOGLE.get_drive_files(_folder_id)
            if "members" in i["title"]
        ][0]
        self.GOOGLE.download_from_drive(_file, ".\data\members.xlsx")
        # open spreadsheet, select correct tab and load into list of dictionaries
        wb = openpyxl.load_workbook(".\data\members.xlsx")
        sheet = wb["Form Responses 2"]
        raw_data = [
            {
                sheet.cell(row=1, column=j).value: sheet.cell(row=i + 1, column=j).value
                for j in range(2, sheet.max_column + 1)
            }
            for i in range(1, sheet.max_row)
        ]

        # clean data and shape data correctly (join placas in single list)
        _wa = list(raw_data[0].keys())[8]
        for r, row in enumerate(copy(raw_data)):
            placas = []
            for col in row:
                if type(row[col]) == float:
                    raw_data[r][col] = str(int(row[col]))
                if type(raw_data[r][col]) == str and "@" in raw_data[r][col]:
                    raw_data[r][col] = raw_data[r][col].lower()
                if "Nombre" in col:
                    raw_data[r][col] = raw_data[r][col].title()
                if "Correo" in col:
                    raw_data[r]["Correo"] = raw_data[r].pop(col)
                if "WhatsApp" in col:
                    raw_data[r][_wa] = 1 if raw_data[r][_wa] == "Sí" else 0
                if "Placa" in col:
                    if raw_data[r][col]:
                        placas.append(
                            raw_data[r][col].replace("-", "").replace(" ", "").upper()
                        )
                    del raw_data[r][col]
            raw_data[r].update(
                {
                    "Placas": placas,
                }
            )
        return raw_data

    def sql(self, table, fields):
        """Create generic SQL statement to insert new records in any table in database."""
        _qmarks = ",".join(["?" for _ in range(len(fields))])
        # return example: REPLACE INTO members (IdMember, Email) VALUES (?,?)
        return f"REPLACE INTO {table} ({','.join(fields)}) VALUES ({_qmarks})"

    def restart_database(self):
        """Erases all data in database. Only for testing purposes. DO NOT USE."""
        SQLSCRIPT1 = os.path.join(os.getcwd(), "static", "sql_script1.sql")
        with open(SQLSCRIPT1, mode="r", encoding="utf-8") as file:
            cmd = file.read()
        self.cursor.executescript(cmd)
