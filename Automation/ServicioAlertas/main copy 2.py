import scrapers
import json, sys, os
import openpyxl
from copy import deepcopy as copy
import time
from gft_utils import ChromeUtils, GoogleUtils
import threading
from datetime import datetime as dt, timedelta as td
import logging
from pprint import pprint
import uuid
import soat, alerts


def start_logger():
    logging.basicConfig(
        filename=LOG_PATH,
        filemode="a",
        format="%(asctime)s | %(levelname)s | %(message)s",
    )
    _log = logging.getLogger(__name__)
    _log.setLevel(logging.INFO)
    return _log


def load_members():
    with open(DATABASE, mode="r", encoding="utf-8") as file:
        return json.load(file)


def save_members(members):
    with open(DATABASE, mode="w+", encoding="utf-8") as file:
        json.dump(members, file, indent=4)
    LOG.info("Database updated sccuesfully")


def load_raw_members():
    # download latest form responses
    G = GoogleUtils()
    _folder_id = "1Az6eM7Fr9MUqNhj-JtvOa6WOhYBg5Qbs"
    _file = [i for i in G.get_drive_files(_folder_id) if "members" in i["title"]][0]
    G.download_from_drive(_file, ".\data\members.xlsx")

    # open latest form responses
    wb = openpyxl.load_workbook(".\data\members.xlsx")
    sheet = wb["Form Responses 2"]
    raw_data = [
        {
            sheet.cell(row=1, column=j).value: sheet.cell(row=i + 1, column=j).value
            for j in range(2, sheet.max_column + 1)
        }
        for i in range(1, sheet.max_row)
    ]

    # clean data and join placas in one list
    for r, row in enumerate(copy(raw_data)):
        placas = []
        for col in row:
            if type(row[col]) == float:
                raw_data[r][col] = str(int(row[col]))
            if type(raw_data[r][col]) == str and "@" in raw_data[r][col]:
                raw_data[r][col] = raw_data[r][col].lower()
            if "Nombre" in col:
                raw_data[r][col] = raw_data[r][col].title()
            if "Correo" in col:
                raw_data[r]["Correo"] = raw_data[r].pop(col)
            if "Placa" in col:
                if raw_data[r][col]:
                    placas.append(
                        raw_data[r][col].replace("-", "").replace(" ", "").upper()
                    )
                del raw_data[r][col]
        raw_data[r].update(
            {
                "Placas": placas,
            }
        )
    return raw_data


def add_new_members(all_members, existing_members):
    correlativo = len(existing_members)
    existing_members_docs = [
        i["Datos"]["Número de Documento"] for i in existing_members
    ]
    for all_member in all_members:
        if str(all_member["Número de Documento"]) not in existing_members_docs:
            _new_record = {
                "Correlativo": correlativo,
                "Codigo": "SAP-" + str(uuid.uuid4())[-6:].upper(),
                "Datos": all_member,
                "Resultados": {
                    "Satimp": [],
                    "Satimp_Actualizado": "01/01/1999",
                    "Brevete": {},
                    "Brevete_Actualizado": "01/01/1999",
                    "Revtec": [],
                    "Revtec_Actualizado": "01/01/1999",
                    "Sutran": [],
                    "Sutran_Actualizado": "01/01/1999",
                    "Soat": [],
                    "Soat_Actualizado": "01/01/1999",
                },
                "Envios": {
                    "Bienvenida": {},
                    "Resumen": [],
                    "Alertas": {
                        "Brevete": [],
                        "Satimp": [],
                        "Revtec": [],
                        "Sutran": [],
                    },
                },
            }

            correlativo += 1
            existing_members.append(_new_record)
            LOG.info(
                f"Appended new record: {_new_record['Correlativo']} - {_new_record['Datos']['Nombre y Apellido']}"
            )

    return existing_members


def get_records_to_process(members):
    # TODO: beyond welcome
    docs_to_process, placas_to_process = [], []
    for member in members[15:16]:
        if not member["Envios"]["Bienvenida"]:
            docs_to_process.append(
                (
                    member["Correlativo"],
                    -1,
                    member["Datos"]["Documento Tipo"],
                    member["Datos"]["Número de Documento"],
                    "",
                )
            )
            for k, placa in enumerate(member["Datos"]["Placas"]):
                placas_to_process.append((member["Correlativo"], k, "", "", placa))

    return docs_to_process, placas_to_process


def gather_soat(members, placas_to_process):

    SOAT = scrapers.Soat()
    scraper_responses = soat.main(SOAT, placas_to_process)

    new_response = [{"Soat": [[], [], []]} for _ in range(len(members))]
    for response in scraper_responses:
        rec, pos, data = response
        new_response[rec]["Soat"][pos] = data

    # delete dictionary keys that have no data to update
    new_responses = [
        {k: v for k, v in j.items() if v and v != [[], [], []]} for j in new_response
    ]

    for k, response in enumerate(new_responses):
        members[k]["Resultados"].update(response)
        members[k]["Resultados"]["Soat_Actualizado"] = dt.now().strftime("%d/%m/%Y")
    return members


def gather(members, docs_to_process, placas_to_process):
    URLS = [
        "https://www.sat.gob.pe/WebSitev8/IncioOV2.aspx",
        "https://licencias.mtc.gob.pe/#/index",
        "https://portal.mtc.gob.pe/reportedgtt/form/frmconsultaplacaitv.aspx",
        "https://www.sutran.gob.pe/consultas/record-de-infracciones/record-de-infracciones/",
    ]
    scraper = [
        (scrapers.Satimp(None), docs_to_process),
        (scrapers.Brevete(None, URLS[1]), docs_to_process),
        (scrapers.Revtec(None), placas_to_process),
        (scrapers.Sutran(None), placas_to_process),
    ]
    threads = []
    for index, (url, scrap) in enumerate(zip(URLS, scraper)):
        _t = threading.Thread(target=full_update, args=(url, scrap[0], scrap[1], index))
        threads.append(_t)
        _t.start()

    for thread in threads:
        thread.join()

    # join responses in member list order
    new_response = [
        {"Satimp": {}, "Brevete": {}, "Revtec": [[], [], []], "Sutran": [[], [], []]}
        for _ in range(len(members))
    ]
    for response, header in zip(responses, HEADERS):
        for resp in response:
            rec, pos, data = resp
            if pos == -1:
                new_response[rec].update({header: data})
            else:
                new_response[rec][header][pos] = data

    # delete dictionary keys that have no data to update
    # TODO: Sutran?
    new_responses = [
        {k: v for k, v in j.items() if v and v != [[], [], []]} for j in new_response
    ]

    for k, response in enumerate(new_responses):
        if response:
            members[k]["Resultados"].update(response)
        for header in HEADERS:
            members[k]["Resultados"][f"{header}_Actualizado"] = dt.now().strftime(
                "%d/%m/%Y"
            )
    return members


def full_update(URL, scraper, data_to_process, index):
    # define Chromedriver and open url for first time
    scraper.WEBD = ChromeUtils().init_driver(
        headless=True, verbose=False, maximized=True
    )
    scraper.WEBD.get(URL)
    time.sleep(3)

    # iterate on all records that require updating
    for k, data in enumerate(data_to_process):
        print(f"{index}: {(k+1)/len(data_to_process)*100:.1f}%")
        # get scraper data
        try:
            attempts, new_record = 0, ""
            while not new_record and attempts < 5:
                new_record, _ = scraper.browser(
                    doc_tipo=data[2], doc_num=data[3], placa=data[4]
                )
                responses[index].append((data[0], data[1], new_record))
                attempts += 1
        except KeyboardInterrupt:
            quit()
        except:
            LOG.warning(f"No proceso {data}")
            responses[index].append((data[0], data[1], ""))


def main():
    # Download raw list of all members from form and add new ones with default data
    if "UPDATE" in sys.argv:
        all_members = load_raw_members()
        existing_members = load_members()
        members = add_new_members(
            all_members=all_members, existing_members=existing_members
        )

        # Run SOAT scrape (if selected) then Automated scrapes (cannot run parallel because of VPN use)
        docs_to_process, placas_to_process = get_records_to_process(members)
        if "SOAT" in sys.argv:
            members = gather_soat(members, placas_to_process)
        members = gather(members, docs_to_process, placas_to_process)

        # Save updated database from scraping
        save_members(members)

    if "ALERT" in sys.argv:
        members = load_members()

        # get list of records to process for each alert
        welcome_list, regular_list, warning_list, timestamps = alerts.get_alert_lists(
            members
        )
        # compose and send alerts
        members = alerts.send_alerts(
            LOG,
            members,
            welcome_list,
            regular_list,
            warning_list,
            EMAIL="EMAIL" in sys.argv,
            timestamps=timestamps,
        )

        # Save updated database with timestamps and unique ids of sent alerts
        # pprint(members)
        save_members(members)


if __name__ == "__main__":
    if not "TEST" in sys.argv:
        DATABASE = os.path.join(os.getcwd(), "data", "members.json")
    else:
        DATABASE = os.path.join(os.getcwd(), "data", "members_test.json")
    HEADERS = ["Satimp", "Brevete", "Revtec", "Sutran"]
    LOG_PATH = os.path.join(os.getcwd(), "logs", "alerts_log.txt")
    LOG = start_logger()
    LOG.info("Start Program.")

    responses = [[] for _ in range(4)]

    main()
    LOG.info("-" * 20 + " End Program " + "-" * 20)

"""
Prouction Sequence:
- python main.py UPDATE SOAT
- pythom main.py ALERT EMAIL

Automated only:
- python main.py UPDATE

Testing Alerts only (no update)
- python main.py ALERT
"""
