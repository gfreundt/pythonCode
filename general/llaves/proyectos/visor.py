from tkinter import END, StringVar
import ttkbootstrap as ttkb
import openpyxl as pyxl
from fpdf import FPDF
from copy import deepcopy as copy

from proyectos.cargar2 import Cargar
from proyectos.editar import Editar


class Visor:

    def __init__(self, cursor, conn):

        self.cursor = cursor
        self.conn = conn

        self.detalle = False

    def gui_pre_cargar(self, previous):

        self.previous = previous

        # abrir dialogo para elegir archivo
        cargar = Cargar(self)
        cargar.gui("proyectos")

    def gui_post_cargar(self):

        # capturar nombre de proyecto y libro que vienen del dialogo para elegir archivo
        self.nombre_proyecto = self.archivo_elegido[0]
        self.libro_origen = self.archivo_elegido[1]

        editar = Editar(self)

        # crear nueva ventana, dimensionar
        self.window = ttkb.Toplevel()
        self.window.geometry(
            f"2000x{int(int(self.window.winfo_screenheight())*.92)}+10+10"
        )

        # definir Frames
        self.frames = {
            "top": ttkb.Frame(self.window),
            "mid": ttkb.Frame(self.window),
            "bottom": ttkb.Frame(self.window, height=100),
        }

        # GUI - Top Frame: botones
        self.frames["top"].pack(pady=10)

        # GUI - Mid Frame: arbol
        self.frames["mid"].pack(pady=10)

        # GUI - Bottom Frame: editar
        self.frames["bottom"].pack(pady=10)

        # definir y colocar botones de menu
        buttons = [
            ttkb.Button(
                self.frames["top"],
                text="Detalle",
                command=lambda: self.menu_detalle(),
            ),
            ttkb.Button(
                self.frames["top"], text="Exportar XLS", command=self.menu_exportar_xls
            ),
            ttkb.Button(
                self.frames["top"], text="Exportar PDF", command=self.menu_exportar_pdf
            ),
            ttkb.Button(
                self.frames["top"],
                text="Editar",
                command=self.menu_editar,
            ),
            ttkb.Button(
                self.frames["top"],
                text="Regresar",
                command=lambda: self.menu_regresar(),
                bootstyle="warning",
            ),
            ttkb.Button(
                self.frames["top"],
                text="Refrescar",
                command=lambda: self.menu_refrescar(),
                bootstyle="success",
            ),
        ]
        for x, button in enumerate(buttons):
            button.grid(row=0, column=x, padx=30, pady=20)

        # crear zona donde se muestra el texto del arbol
        self.area_texto = ttkb.Text(self.frames["mid"], height=63, width=250)
        self.area_texto.pack()

        # genera el texto del arbol al visor y mostrar
        self.menu_detalle()

        editar.gui(frame=self.frames["bottom"])

    def menu_detalle(self):

        arbol = self.genera_texto_arbol()
        self.muestra_arbol(arbol)
        self.detalle = not self.detalle

    def menu_refrescar(self):

        arbol = self.genera_texto_arbol()
        self.muestra_arbol(arbol)

    def menu_exportar_xls(self):
        wb = pyxl.Workbook()
        ws = wb.active

        # crear hoja "Estructura"
        ws.title = "Estructura"
        titulos = [
            "Codigo",
            "GGMK",
            "Formato",
            "Nombre",
            "Notas",
            "Creacion",
            "GMKs",
            "MKs",
            "SMKs",
            "Ks",
        ]
        self.cursor.execute(
            f"SELECT * FROM proyectos WHERE Codigo='{self.nombre_proyecto}'"
        )
        data = self.cursor.fetchone()
        for i, (a, b) in enumerate(zip(titulos, data), start=1):
            ws[f"A{i}"] = str(a)
            ws[f"B{i}"] = str(b)

        # crear hoja "Resumen"
        wb.create_sheet("Resumen")
        ws = wb["Resumen"]
        arbol = self.genera_texto_arbol(detalle=False)
        fila = 0
        for linea in arbol.split("\n"):
            fila += 1
            if "GGMK" in linea:
                ws[f"A{fila}"] = linea
            elif "GMK-" in linea:
                ws[f"B{fila}"] = linea
            elif "MK-" in linea:
                ws[f"C{fila}"] = linea
            elif "K-" in linea:
                ws[f"D{fila}"] = linea
            else:
                fila -= 1

        # crear hoja "Detalle"
        wb.create_sheet("Detalle")
        ws = wb["Detalle"]
        arbol = self.genera_texto_arbol(detalle=True)
        fila = 0
        for linea in arbol.split("\n"):
            fila += 1
            if "GGMK" in linea:
                ws[f"A{fila}"] = linea
            elif "GMK-" in linea:
                ws[f"B{fila}"] = linea
            elif "MK-" in linea:
                ws[f"C{fila}"] = linea
            elif "K-" in linea:
                ws[f"D{fila}"] = linea
            else:
                fila -= 1

        # crear plantilla para carga de proyectos
        wb.create_sheet("PlantillaProyecto")

        # guardar hoja
        wb.save(f"{self.nombre_proyecto}.xlsx")

    def menu_exportar_pdf(self):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        arbol = self.genera_texto_arbol(detalle=True)
        for linea in arbol.split("\n"):
            pdf.cell(200, 10, txt=linea, ln=1, align="L")
        pdf.output("mygfg.pdf")

    def menu_editar(self):

        for widget in self.frames["bottom"].winfo_children():
            widget.config(state="normal")

    def menu_regresar(self):

        self.previous.window.deiconify()
        self.window.destroy()

    def genera_texto_arbol(self, detalle=None):

        # si no obliga un formato de detalle, usar el activo
        if not detalle:
            detalle = self.detalle

        totales = {"gmk": 0, "mk": 0}

        self.cursor.execute(
            f"""SELECT  Secuencia, Jerarquia, CodigoLlave, Nombre, Copias, CodigoPuerta, TipoPuerta,
                        TipoCerradura, Zona1, Zona2, Zona3, Zona4, ZonaCodigo
                        FROM '{self.nombre_proyecto}'"""
        )

        data = [[j if j else "" for j in i] for i in self.cursor.fetchall()]
        output = ""

        for (
            sec,
            jer,
            cod,
            nom,
            cop,
            codp,
            tipp,
            tipc,
            zon1,
            zon2,
            zon3,
            zon4,
            zonc,
        ) in data:

            if jer == "GGMK":
                output += f"{sec} | Codigo:{cod} | Nombre: {nom} | Copias: {cop} |\n"

            if jer == "GMK":
                output += "|\n"
                output += f"|{'-'*8} {sec} | Codigo:{cod} | Nombre: {nom} | Copias: {cop} | Zona: {zon1} {zon2} {zon3} {zon4} {zonc} |\n"
                totales["gmk"] += 1

            if jer == "MK":
                output += f"|{' '*9}|\n"
                output += f"|{' '*9}|{'-'*7} {sec} | Codigo:{cod} | Nombre: {nom} | Copias: {cop} | Zona: {zon1} {zon2} {zon3} {zon4} {zonc} |\n"
                output += f"|{' '*9}|{' '*10}|\n"
                totales["mk"] += 1

                if not detalle:
                    self.cursor.execute(
                        f"SELECT * FROM '{self.nombre_proyecto}' AS T1 JOIN '{self.libro_origen}' AS T2 ON T1.Secuencia = T2.Secuencia WHERE MK = '{cod}'"
                    )
                    _unicas = len(self.cursor.fetchall())
                    output += f"|{' '*9}|{' '*10}|- K-Únicas: {_unicas}\n"

            if jer == "K" and detalle:
                output += f"""|{' '*9}|{' '*10}|- {sec} | Codigo:{cod} | Nombre: {nom} | Copias: {cop} | Zona: {zon1} {zon2} {zon3} {zon4} {zonc} | Puerta: {codp} - {tipp} | Cerradura: {tipc} |\n"""

        # agregar resumen al inicio del texto
        self.cursor.execute(
            f"SELECT * FROM 'proyectos' WHERE Codigo = '{self.nombre_proyecto}'"
        )
        d = self.cursor.fetchone()

        output = (
            f"""{'-'*50}\nProyecto: {self.nombre_proyecto}\nNombre: {d[3]}\nFecha de Creacion: {d[5]}\nTotal GMKs: {int(d[6]):,}\nTotal MKs: {int(d[7]):,}\nTotal Ks: {int(d[9]):,}\n{'-'*50}\n\n"""
            + output
        )

        return output

    def muestra_arbol(self, arbol):
        self.area_texto.delete("1.0", "end")
        self.area_texto.insert(END, arbol)
