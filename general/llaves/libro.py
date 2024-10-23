from random import randrange, shuffle
from copy import deepcopy as copy
from tkinter import Tk, Label, Text, END, font, PhotoImage, Button
from pprint import pprint
from datetime import datetime as dt
import openpyxl as pyxl
from visor import Visor
import libro_crea
from itertools import product


class Libro:

    def __init__(self, conn):

        # conecta a la base de datos
        self.conn = conn
        self.cursor = self.conn.cursor()

        # inicializa visor
        self.vista = Visor(configuracion="libro", proceso=self)

        # variables de estado
        self.detalle = False
        self.status_graba_db = False

    def menu_detalle(self):
        self.detalle = not self.detalle
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def menu_rehacer(self):
        self.text_area.delete("1.0", "end")
        self.text_area.insert(END, "Procesando...\n")
        self.crea_libro()

    def menu_exportar(self):
        wb = pyxl.Workbook()
        ws = wb.active

        arbol = self.genera_texto_arbol(detalle=True)
        fila = 0
        for linea in arbol.split("\n"):
            fila += 1
            if "GGMK" in linea:
                ws[f"A{fila}"] = linea
            elif "GMK-" in linea:
                ws[f"B{fila}"] = linea
            elif "MK-" in linea:
                ws[f"C{fila}"] = linea
            elif "K-" in linea:
                ws[f"D{fila}"] = linea
            else:
                fila -= 1

        wb.save("export.xlsx")

    def menu_guardar(self):
        self.conn.commit()
        self.status_graba_db = True

    def menu_regresar(self):
        if not self.status_graba_db:
            self.cursor.execute(f"DROP TABLE '{self.nombre_tabla}'")
            self.conn.commit()
        self.window.destroy()

    def crea_libro(self, **kwargs):

        # en primera solicitud de libro nuevo, recibe toda esta informacion (no en rehacer libro)
        if kwargs:
            self.formato = kwargs["formato"]
            self.codigo_ggmk = kwargs["codigo_ggmk"]
            self.libro_nombre = kwargs["libro_nombre"]
            self.libro_notas = kwargs["libro_notas"]

        # genera arbol nuevo si el arbol creado no tenga suficientes llaves validas
        arbol = ()
        while len(arbol) < 10:
            arbol, self.nombre_tabla = libro_crea.generador(
                self.codigo_ggmk, self.formato
            )

        # arbol es valido, crea tabla y carga las llaves
        self.cursor.execute(
            f"CREATE TABLE '{self.nombre_tabla}' (GGMK, GMK, MK, SMK, K, Secuencia, Cilindro, MP)"
        )
        self.cursor.executemany(
            f"INSERT INTO '{self.nombre_tabla}' (GGMK, GMK, MK, SMK, K, Secuencia, Cilindro, MP) VALUES (?,?,?,?,?,?,?,?)",
            arbol,
        )
        self.conn.commit()

        print(self.valida_libro_completo())

        # muestra las llaves en el GUI
        self.muestra_arbol()

    def valida_libro_completo(self):

        # extraer toda las filas de la tabla
        self.cursor.execute(f"SELECT * FROM '{self.nombre_tabla}'")
        table_data = self.cursor.fetchall()

        # extrae todos los codigos de las master keys en la tabla
        self.cursor.execute(
            f"SELECT * FROM (SELECT GGMK FROM '{self.nombre_tabla}' UNION SELECT GMK FROM '{self.nombre_tabla}' UNION SELECT MK FROM '{self.nombre_tabla}' UNION SELECT SMK FROM '{self.nombre_tabla}' WHERE GGMK IS NOT 0)"
        )
        todas_maestras = [i[0] for i in self.cursor.fetchall()]

        # extraer todos los codigos de las llaves de la tabla
        self.cursor.execute(f"SELECT K FROM '{self.nombre_tabla}'")
        llaves_usadas = [i[0] for i in self.cursor.fetchall()]

        for row in table_data:
            if not all(
                [
                    valida_llave_es_unica(llave=row[4], llaves_usadas=llaves_usadas),
                    valida_llave_abre_cilindro(
                        llaves=[row[0], row[1], row[2]], cilindro=row[6]
                    ),
                    valida_llave_no_cruzada(
                        cilindro=row[6],
                        mis_maestras=[i for i in row[:5] if i != 0],
                        todas_maestras=todas_maestras,
                    ),
                ]
            ):
                return False

        return True

    def cargar_proyecto(self, tabla):
        self.nombre_tabla = tabla
        self.vista = Visor(configuracion="proyecto", proceso=self)
        self.vista.mostrar()

    def carga_libro(self, tabla):

        self.nombre_tabla = tabla
        self.vista = Visor(configuracion="libro", proceso=self)
        self.vista.mostrar()

    def muestra_arbol(self):
        self.vista.mostrar()


def valida_llave_abre_cilindro(llaves, cilindro):

    # crea lista de todas las llaves que pueden abrir ese cilindro
    opciones = [
        f"{pos[1]}{int(pos[1]) + int(pos[3])}" if ":" in pos else pos[1]
        for pos in cilindro.split("]")[:-1]
    ]
    g = [i.replace("[", "").replace(":", "") for i in opciones]
    cilindros = ["".join(i) for i in product(*g)]

    # revisa llaves de lista y ver si abren cilindro
    for llave in llaves:
        if llave not in cilindros:
            return False
    return True


def valida_llave_es_unica(llave, llaves_usadas):

    # revisa que solo haya una vez este codigo de llave en todo el libro
    if llaves_usadas.count(llave) == 1:
        return True
    return False


def valida_llave_no_cruzada(cilindro, mis_maestras, todas_maestras):

    # revisa cilindro contra todas las llaves maestras que no le corresponden para asegurar que no lo abran
    for maestra in todas_maestras:
        if maestra in mis_maestras:
            continue
        if valida_llave_abre_cilindro([maestra], cilindro):
            return False
    return True
