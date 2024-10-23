from copy import deepcopy as copy
from tkinter import Tk, Text, END, Button
from pprint import pprint
from datetime import datetime as dt
import uuid
import openpyxl as pyxl


class Visor:

    def __init__(self, configuracion, proceso):
        self.configuracion = configuracion
        self.proceso = proceso
        self.fecha = dt.strftime(dt.now(), "%Y-%m-%d")
        self.file_name = f"L{str(uuid.uuid4())[-10:]}.json"
        self.detalle = False
        self.status_graba_db = False

    def mostrar(self):

        # crear nueva ventana, dimensionar
        self.window = Tk()
        self.window.geometry("1000x1300")

        # definir y colocar botones de menu
        self.buttons = [
            (Button(self.window, text="Detalle", command=self.menu_detalle), (75, 20)),
            (
                Button(
                    self.window, text="Exportar XLS", command=self.menu_exportar_xls
                ),
                (150, 20),
            ),
            (
                Button(
                    self.window, text="Exportar PDF", command=self.menu_exportar_pdf
                ),
                (250, 20),
            ),
            (Button(self.window, text="Guardar", command=self.menu_guardar), (350, 20)),
            (
                Button(self.window, text="Regresar", command=self.menu_regresar),
                (420, 20),
            ),
        ]
        for button, pos in self.buttons:
            button.place(x=pos[0], y=pos[1])

        # crear zona donde se muestra el texto del arbol
        self.text_area = Text(self.window, height=100, width=130)
        self.text_area.place(x=10, y=60)

        # genera el texto del arbol al visor y mostrar
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def menu_detalle(self):
        self.detalle = not self.detalle
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def menu_rehacer(self):
        self.crea_libro(codigo_ggmk=copy(self.ggmk["codigo"]))
        self.detalle = False
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def menu_exportar_xls(self):
        wb = pyxl.Workbook()
        ws = wb.active

        arbol = self.genera_texto_arbol(detalle=True)
        fila = 0
        for linea in arbol.split("\n"):
            fila += 1
            if "GGMK" in linea:
                ws[f"A{fila}"] = linea
            elif "GMK-" in linea:
                ws[f"B{fila}"] = linea
            elif "MK-" in linea:
                ws[f"C{fila}"] = linea
            elif "K-" in linea:
                ws[f"D{fila}"] = linea
            else:
                fila -= 1

        wb.save(f"{self.proceso.nombre_tabla}.xlsx")

    def menu_exportar_pdf(self):
        return

    def menu_guardar(self):
        self.proceso.conn.commit()
        self.status_graba_db = True

    def menu_regresar(self):
        if not self.status_graba_db:
            self.proceso.cursor.execute(f"DROP TABLE '{self.proceso.nombre_tabla}'")
            self.proceso.conn.commit()
        self.window.destroy()

    def genera_texto_arbol(self, detalle):

        if "proyecto" in self.configuracion:
            return self.genera_texto_arbol_proyecto(detalle=detalle)
        else:
            return self.genera_texto_arbol_libro(detalle=detalle)

    def genera_texto_arbol_proyecto(self, detalle):

        self.proceso.cursor.execute(f"SELECT * FROM '{self.proceso.nombre_proyecto}'")
        data = self.proceso.cursor.fetchall()

        previous = [0] * 10
        output = f"GGMK <> Codigo:{data[0][0]}\n"

        for line, d in enumerate(data):

            gmk = d[1]
            mk = d[2]
            k = d[4]
            sec = d[5]
            cil = d[6]
            mp = d[7]
            tpuerta = d[8]
            cerrad = d[9]
            cop = d[10]
            nombre = " - ".join([str(i) for i in d[11:16] if i])

            if gmk != previous[1]:
                output += "|\n"
                output += f"|{'-'*8} GM{sec[:4]} <> Codigo: {gmk}\n"

            if mk != previous[2]:
                output += f"|{' '*9}|\n"
                output += f"|{' '*9}|{'-'*7} M{sec[:7]} <> Codigo:{mk}\n"
                output += f"|{' '*9}|{' '*8}|\n"

                if not detalle:
                    self.proceso.cursor.execute(
                        f"SELECT * FROM '{self.proceso.nombre_proyecto}' WHERE MK = '{mk}'"
                    )
                    _unicas = len(self.proceso.cursor.fetchall())
                    output += f"|{' '*9}|{' '*8}K-Unicas: {_unicas}\n"

            if detalle:
                output += f"|{' '*9}|{' '*8}|- {sec} <> Codigo:{k} - Cilindro ({mp} MP): {cil:<30} - Copias: {cop} [{nombre if nombre else '(sin nombre)'}]\n"

            previous = copy(data[line])

        return output

    def genera_texto_arbol_libro(self, detalle):

        self.proceso.cursor.execute(f"SELECT * FROM '{self.proceso.nombre_tabla}'")
        data = self.proceso.cursor.fetchall()

        previous = [0] * 10

        output = f"{'-'*50}\nLibro: {self.proceso.nombre_tabla}\nValidaciones: OK\nCilindros: {len(data):,}\n{'-'*50}\n"

        output += f"GGMK <> Codigo:{data[0][0]}\n"

        for line, d in enumerate(data):

            gmk = d[1]
            mk = d[2]
            k = d[4]
            sec = d[5]
            cil = d[6]
            mp = d[7]

            if gmk != previous[1]:
                output += "|\n"
                output += f"|{'-'*8} GM{sec[:4]} <> Codigo: {gmk}\n"

            if mk != previous[2]:
                output += f"|{' '*9}|\n"
                output += f"|{' '*9}|{'-'*7} M{sec[:7]} <> Codigo:{mk}\n"
                output += f"|{' '*9}|{' '*8}|\n"

                if not detalle:
                    self.proceso.cursor.execute(
                        f"SELECT * FROM '{self.proceso.nombre_tabla}' WHERE MK = '{mk}'"
                    )
                    _unicas = len(self.proceso.cursor.fetchall())
                    output += f"|{' '*9}|{' '*8}K-Unicas: {_unicas}\n"

            if detalle:
                output += f"|{' '*9}|{' '*8}|- {sec} <> Codigo:{k} - Cilindro ({mp} MP): {cil:<30} -\n"

            previous = copy(data[line])

        return output

    def muestra_arbol(self, arbol):
        self.text_area.delete("1.0", "end")
        self.text_area.insert(END, arbol)
