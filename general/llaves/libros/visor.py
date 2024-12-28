from tkinter import END
import ttkbootstrap as ttkb
import openpyxl as pyxl
from openpyxl.styles import Alignment, Color, Fill
from openpyxl.cell import Cell

from fpdf import FPDF
from copy import deepcopy as copy


def mostrar(cursor, nombre_tabla, main_window):

    global detalle
    detalle = False

    # crear nueva ventana, dimensionar
    window = ttkb.Toplevel()
    window.geometry("1300x2000+10+10")

    # GUI - Top Frame: botones
    top_frame = ttkb.Frame(window)
    top_frame.pack(pady=10)

    # GUI - Bottom Frame: arbol
    bottom_frame = ttkb.Frame(window)
    bottom_frame.pack(pady=10)

    # definir y colocar botones de menu
    buttons = [
        ttkb.Button(
            top_frame,
            text="Detalle",
            command=lambda: menu_detalle(
                cursor=cursor,
                nombre_tabla=nombre_tabla,
                area_texto=area_texto,
            ),
        ),
        ttkb.Button(
            top_frame,
            text="Exportar XLS",
            command=lambda: menu_exportar_xls(cursor, nombre_tabla),
        ),
        ttkb.Button(top_frame, text="Exportar PDF", command=menu_exportar_pdf),
        ttkb.Button(
            top_frame,
            text="Regresar",
            command=lambda: menu_regresar(main_window=main_window, window=window),
            bootstyle="warning",
        ),
    ]
    for x, button in enumerate(buttons):
        button.grid(row=0, column=x, padx=30, pady=20)

    # crear zona donde se muestra el texto del arbol
    area_texto = ttkb.Text(bottom_frame, height=78, width=130)
    area_texto.pack()

    # genera el texto del arbol al visor y mostrar
    menu_detalle(cursor=cursor, nombre_tabla=nombre_tabla, area_texto=area_texto)


def menu_detalle(**kwargs):

    global detalle

    arbol = genera_texto_arbol(
        detalle=detalle,
        cursor=kwargs["cursor"],
        nombre_libro=kwargs["nombre_tabla"],
    )
    muestra_arbol(arbol, area_texto=kwargs["area_texto"])
    detalle = not detalle


def menu_exportar_xls(cursor, nombre_tabla):
    wb = pyxl.Workbook()
    ws = wb.active

    # crear hoja "Estructura"
    ws.title = "Estructura"
    titulos = [
        "Codigo",
        "GGMK",
        "Formato",
        "Nombre",
        "Notas",
        "Creacion",
        "GMKs",
        "MKs",
        "SMKs",
        "Ks",
    ]
    cursor.execute(f"SELECT * FROM libros WHERE Codigo='{nombre_tabla}'")
    data = cursor.fetchone()
    for i, (a, b) in enumerate(zip(titulos, data), start=1):
        ws[f"A{i}"] = str(a)
        ws[f"B{i}"] = str(b)

    # crear hoja "Resumen"
    wb.create_sheet("Arbol Resumen")
    ws = wb["Arbol Resumen"]
    arbol = genera_texto_arbol(cursor, nombre_tabla, detalle=False)
    fila = 0
    for linea in arbol.split("\n"):
        fila += 1
        if "GGMK" in linea:
            ws[f"A{fila}"] = linea
        elif "GMK-" in linea:
            ws[f"B{fila}"] = linea
        elif "MK-" in linea:
            ws[f"C{fila}"] = linea
        elif "K-" in linea:
            ws[f"D{fila}"] = linea
        else:
            fila -= 1

    # crear hoja "Detalle"
    wb.create_sheet("Arbol Detalle")
    ws = wb["Arbol Detalle"]
    arbol = genera_texto_arbol(cursor, nombre_tabla, detalle=True)
    fila = 0
    for linea in arbol.split("\n"):
        # linea = linea.replace("|", "").replace("-", "")
        fila += 1
        if "GGMK" in linea:
            ws[f"A{fila}"] = linea
        elif "GMK-" in linea:
            ws[f"B{fila}"] = linea
        elif "MK-" in linea:
            ws[f"C{fila}"] = linea
        elif "K-" in linea:
            ws[f"D{fila}"] = linea
        else:
            fila -= 1

    # crear hoja "Puertas"
    wb.create_sheet("Puertas")
    ws = wb["Puertas"]
    cursor.execute(f"SELECT * FROM '{nombre_tabla}'")
    titulos = [
        (
            "GGMK",
            "GMK",
            "MK",
            "SMK",
            "K",
            "Secuencia",
            "Cilindro",
            "MP",
            "AsignadaProyecto",
        )
    ]
    data = cursor.fetchall()

    for i, (a, b, c, d, e, f, g, h, j) in enumerate(titulos + data, start=1):
        ws[f"A{i}"] = str(a)
        ws[f"B{i}"] = str(b)
        ws[f"C{i}"] = str(c)
        ws[f"D{i}"] = str(d)
        ws[f"E{i}"] = str(e)
        ws[f"F{i}"] = str(f)
        ws[f"G{i}"] = str(g)
        ws[f"H{i}"] = str(h)
        ws[f"I{i}"] = str(j)

    # crear plantilla para carga de proyectos
    wb.create_sheet("PlantillaProyecto")
    ws = wb["PlantillaProyecto"]

    previous_b = ""
    previous_a = ""

    b_splits = []
    a_splits = []

    prev_a_split = 0
    prev_b_split = 0

    color_index0 = color_index1 = color_index2 = -1
    outer_colors, inner_colors = set_colors()

    for i, key_code in enumerate(data, start=5):

        _, a, b, c = key_code[5].split("-")

        if b != previous_b:
            b_splits.append((prev_b_split, i - 1))
            prev_b_split = i
            color_index1 += 1
        previous_b = str(b)

        if a != previous_a:
            a_splits.append((prev_a_split, i - 1))
            prev_a_split = i
            color_index0 += 1
        previous_a = str(a)

        cell = ws[f"A{i}"]
        cell.value = str(a)
        cell.fill = pyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=outer_colors[color_index0]
        )

        cell = ws[f"B{i}"]
        cell.value = str(b)
        cell.fill = pyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=inner_colors[color_index1]
        )
        cell = ws[f"C{i}"]
        cell.value = str(c)
        cell.fill = pyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=inner_colors[color_index1]
        )

        cell = ws[f"D{i}"]
        cell.value = str(key_code[5])
        cell.fill = pyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=inner_colors[color_index1]
        )

    for s in a_splits[1:]:
        ws.merge_cells(start_row=s[0], start_column=1, end_row=s[1], end_column=1)

    for s in b_splits[1:]:
        ws.merge_cells(start_row=s[0], start_column=2, end_row=s[1], end_column=2)

    # formatear y guardar hoja
    for cell in ws["A:A"] + ws["B:B"]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in (("A", 5), ("B", 10), ("C", 15), ("D", 30)):
        ws.column_dimensions[col].width = width
    wb.save(f"{nombre_tabla}.xlsx")


def menu_exportar_pdf(cursor, nombre_libro):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    arbol = genera_texto_arbol(detalle, cursor, nombre_libro)
    for linea in arbol.split("\n"):
        pdf.cell(200, 10, txt=linea, ln=1, align="L")
    pdf.output("mygfg.pdf")


def menu_regresar(window, main_window):

    main_window.deiconify()
    window.destroy()


def genera_texto_arbol(cursor, nombre_libro, detalle):

    # extrae data de libro de base de datos
    cursor.execute(f"SELECT * FROM '{nombre_libro}'")
    data = cursor.fetchall()

    # variables en blanco para iniciar proceso
    previous = [0] * 10
    totales = {"gmk": 0, "mk": 0}

    # ingresar cabecera de texto
    output = f"GGMK <> Codigo:{data[0][0]}\n"

    # loopear todas las llaves y armar el arbol
    for line, (_, gmk, mk, _, k, sec, cil, mp, _) in enumerate(data):

        if gmk != previous[1]:
            output += "|\n"
            output += f"|{'-'*8} GM{sec[:4]} <> Codigo: {gmk}\n"
            totales["gmk"] += 1

        if mk != previous[2]:
            output += f"|{' '*9}|\n"
            output += f"|{' '*9}|{'-'*7} M{sec[:8]} <> Codigo:{mk}\n"
            output += f"|{' '*9}|{' '*8}|\n"
            totales["mk"] += 1

            if not detalle:
                cursor.execute(f"SELECT * FROM '{nombre_libro}' WHERE MK = '{mk}'")
                _unicas = len(cursor.fetchall())
                output += f"|{' '*9}|{' '*8}K-Unicas: {_unicas}\n"

        if detalle:
            output += f"|{' '*9}|{' '*8}|- {sec} <> Codigo:{k} - Cilindro ({mp} MP): {cil:<30} -\n"

        previous = copy(data[line])

    # agregar resumen al inicio del texto
    output = (
        f"""{'-'*50}\nLibro: {nombre_libro}\nValidaciones: OK\nTotal GMKs: {totales['gmk']:,}\nTotal MKs: {totales['mk']:,}\nTotal Ks: {len(data):,}\n{'-'*50}\n"""
        + output
    )

    return output


def muestra_arbol(arbol, area_texto):
    area_texto.delete("1.0", "end")
    area_texto.insert(END, arbol)


def set_colors():
    c = [
        "ff0000",
        "0000ff",
        "00ff00",
        "ffa500",
        "ffffff",
        "000000",
        "ffff00",
        "800080",
        "c0c0c0",
        "a52a2a",
        "808080",
        "ffc0cb",
        "808000",
        "800000",
        "ee82ee",
        "36454f",
        "ff00ff",
        "cd7f32",
        "fffdd0",
        "ffd700",
        "d2b48c",
        "008080",
        "ffdb58",
        "000080",
        "ff7f50",
        "800020",
        "e6e6fa",
        "e0b0ff",
        "e0f7fa",
        "ffe5b4",
        "b7410e",
        "4b0082",
        "e0115f",
        "32cd32",
        "fa8072",
        "007fff",
        "f5f5dc",
        "996666",
        "40e0d0",
        "00ffff",
        "3eb489",
        "87ceeb",
        "dc143c",
        "f4c430",
        "fff44f",
        "43254f",
        "ff00ff",
        "ffbf00",
        "2e8b57",
        "006400",
        "eae0c8",
        "fffff0",
        "f28500",
        "733635",
        "de3163",
        "50c878",
        "664238",
        "0f52ba",
        "c8a2c8",
        "65000b",
        "0000ff",
        "808080",
        "C0A392",
        "6f4e37",
        "0a0a0a",
        "00ff00",
        "635147",
    ]

    return c[:5], c[6:]
