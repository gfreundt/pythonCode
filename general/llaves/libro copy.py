from random import randrange, shuffle
from copy import deepcopy as copy
from tkinter import Tk, Label, Text, END, font, PhotoImage, Button
from pprint import pprint
from datetime import datetime as dt
import json, time
import uuid
import openpyxl as pyxl


def crea_cilindro(ggmk, key):
    ggmk = (int(i) for i in ggmk)
    key = (int(i) for i in key)

    cilindro = []
    for g, k in zip(ggmk, key):
        if g == k:
            cilindro.append(f"[{g}]")
        else:
            cilindro.append(f"[{min(g,k)}:{abs(g-k)}]")

    return "".join(cilindro)


def valida_codigo(codigo):

    # string to tuple
    codigo = tuple(int(i) for i in codigo)

    # no puede haber 8 o 9 de diferencia entre pines
    for position in range(len(codigo) - 1):
        c = codigo[position]
        n = codigo[position + 1]
        if abs(n - c) >= 8:
            return False

    # no puede haber una secuencia [par-impar-par-impar-par-impar] o [impar-par-impar-par-impar-par]
    test = [int(i) % 2 for i in codigo]
    if test == [0, 1, 0, 1, 0, 1] or test == [1, 0, 1, 0, 1, 0]:
        return False

    # no pueden haber tres pines seguidos iguales
    for pos in range(4):
        if codigo[pos : pos + 3].count(codigo[pos]) == 3:
            return False

    return True


def valida_llave_abre_cilindro(llave, cilindro):
    opciones = []
    for pos in cilindro.split("]")[:-1]:
        if ":" in pos:
            opciones.append((pos[1], int(pos[1]) + int(pos[3])))
        else:
            opciones.append((pos[1],))

    cilindros = []
    for a in opciones[0]:
        for b in opciones[1]:
            for c in opciones[2]:
                for d in opciones[3]:
                    for e in opciones[4]:
                        for f in opciones[5]:
                            cilindros.append(f"{a}{b}{c}{d}{e}{f}")

    if llave in cilindros:
        return 1

    return 0


def cantidad_master_pines(cilindro):
    return cilindro.count(":")


class Libro:

    def __init__(self, libro, notas):
        self.libro = libro
        self.notas = notas
        self.fecha = dt.strftime(dt.now(), "%Y-%m-%d")
        self.file_name = f"L{str(uuid.uuid4())[-10:]}.json"
        self.detalle = False

        self.window = Tk()
        self.window.geometry("800x1300")
        self.text_area = Text(self.window, height=100, width=80)
        self.text_area.place(x=10, y=60)

        # definir y colocar botones de menu
        self.buttons = [
            Button(self.window, text="Detalle", command=self.libro_detalle),
            Button(self.window, text="Rehacer", command=self.libro_rehacer),
            Button(self.window, text="Exportar", command=self.libro_exportar),
            Button(self.window, text="Guardar", command=self.libro_guardar),
            Button(self.window, text="Regresar", command=self.arbol_regresar),
        ]
        for x, button in enumerate(self.buttons, start=1):
            button.place(x=x * 75, y=20)

    def libro_detalle(self):
        self.detalle = not self.detalle
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def libro_rehacer(self):
        self.crea_libro(codigo_ggmk=copy(self.ggmk["codigo"]))
        self.detalle = False
        arbol = self.genera_texto_arbol(detalle=self.detalle)
        self.muestra_arbol(arbol)

    def libro_exportar(self):
        wb = pyxl.Workbook()
        ws = wb.active

        arbol = self.genera_texto_arbol(detalle=True)
        fila = 0
        for linea in arbol.split("\n"):
            fila += 1
            if "GGMK" in linea:
                ws[f"A{fila}"] = linea
            elif "GMK-" in linea:
                ws[f"B{fila}"] = linea
            elif "MK-" in linea:
                ws[f"C{fila}"] = linea
            elif "K-" in linea:
                ws[f"D{fila}"] = linea
            else:
                fila -= 1

        wb.save("export.xlsx")

    def libro_guardar(self):
        with open(self.file_name, "w") as outfile:
            json_object = json.dumps(self.ggmk, indent=4)
            outfile.write(json_object)

    def arbol_regresar(self):
        self.window.destroy()

    def crea_libro(self, codigo_ggmk):
        self.todas_maestras = []
        self.genera_mp_pos()

        self.ggmk = {
            "codigo": copy(codigo_ggmk),
            "nombre": "Llave GGMK",
            "libro": self.libro,
            "notas": self.notas,
            "fecha": dt.strftime(dt.now(), "%Y-%m-%d"),
            "subkeys": [],
        }

        self.llaves_usadas = [copy(codigo_ggmk)]
        self.todas_maestras.append(copy(codigo_ggmk))

        c1 = 0
        while True:
            x1 = self.siguiente_codigo(
                tipo="gmk",
                codigo_base=(
                    self.ggmk["codigo"]
                    if not self.ggmk["subkeys"]
                    else self.ggmk["subkeys"][-1]["codigo"]
                ),
            )
            if x1 == -1:
                break
            self.ggmk["subkeys"].append(
                {
                    "codigo": copy(x1),
                    "secuencia": f"GMK-{c1+1:02d}",
                    "nombre": f"Llave {c1+1:02d}",
                    "subkeys": [],
                }
            )
            self.llaves_usadas.append(copy(x1))
            self.todas_maestras.append(copy(x1))

            c2 = 0
            while True:
                x2 = self.siguiente_codigo(
                    tipo="mk",
                    codigo_base=(
                        x1
                        if not self.ggmk["subkeys"][c1]["subkeys"]
                        else self.ggmk["subkeys"][c1]["subkeys"][-1]["codigo"]
                    ),
                )
                if x2 == -1:
                    break
                self.ggmk["subkeys"][c1]["subkeys"].append(
                    {
                        "codigo": copy(x2),
                        "secuencia": f"MK-{c1+1:02d}-{c2+1:02d}",
                        "nombre": f"Llave {c1+1:02d}-{c2+1:02d}",
                        "subkeys": [],
                    }
                )
                self.llaves_usadas.append(copy(x2))
                self.todas_maestras.append(copy(x2))

                c3 = 0
                while True:
                    x3 = self.siguiente_codigo(
                        tipo="k",
                        codigo_base=(
                            x2
                            if not self.ggmk["subkeys"][c1]["subkeys"][c2]["subkeys"]
                            else self.ggmk["subkeys"][c1]["subkeys"][c2]["subkeys"][-1][
                                "codigo"
                            ]
                        ),
                    )
                    if x3 == -1:
                        break
                    self.ggmk["subkeys"][c1]["subkeys"][c2]["subkeys"].append(
                        {
                            "codigo": copy(x3),
                            "cilindro": crea_cilindro(ggmk=codigo_ggmk, key=copy(x3)),
                            "secuencia": f"K-{c1+1:02d}-{c2+1:02d}-{c3+1:03d}",
                            "nombre": f"Llave {c1+1:02d}-{c2+1:02d}-{c3+1:03d}",
                        }
                    )
                    self.llaves_usadas.append(copy(x3))
                    c3 += 1
                c2 += 1
            c1 += 1

    def siguiente_codigo(self, tipo, codigo_base):

        while True:

            codigo_nuevo = [i for i in codigo_base]
            pos = self.mp_pos[tipo]

            if len(pos) == 1:

                codigo_nuevo[pos[0]] = str((int(codigo_nuevo[pos[0]]) + 2) % 10)
                codigo_nuevo = "".join(codigo_nuevo)

                if codigo_nuevo in self.llaves_usadas:
                    return -1

                if valida_codigo(codigo_nuevo):
                    return codigo_nuevo

                return -1

            else:

                for a in range(0, 9, 2):
                    codigo_nuevo[pos[3]] = str((int(codigo_nuevo[pos[3]]) + a) % 10)
                    for b in range(0, 9, 2):
                        codigo_nuevo[pos[2]] = str((int(codigo_nuevo[pos[2]]) + b) % 10)
                        for c in range(0, 9, 2):
                            codigo_nuevo[pos[1]] = str(
                                (int(codigo_nuevo[pos[1]]) + c) % 10
                            )
                            for d in range(0, 9, 2):
                                codigo_nuevo[pos[0]] = str(
                                    (int(codigo_nuevo[pos[0]]) + d) % 10
                                )
                                cn = "".join(codigo_nuevo)
                                if cn not in self.llaves_usadas and valida_codigo(cn):
                                    return cn
                return -1

    def genera_mp_pos(self):
        _pines = list(range(0, 6))
        shuffle(_pines)

        # formato 1-1-1-4
        self.mp_pos = {"gmk": _pines[0:1], "mk": _pines[1:2], "k": _pines[2:]}

        print(self.mp_pos)

    def genera_texto_arbol(self, detalle):
        totalx = 0
        total = 1

        output = f"GGMK | Codigo:{self.ggmk['codigo']}\n"
        for p, gmk in enumerate(self.ggmk["subkeys"], start=1):
            total += 1
            output += "|\n"
            output += f"|{'-'*9}GMK-{p:02d} <> Codigo:{gmk['codigo']}\n"

            for q, mk in enumerate(gmk["subkeys"], start=1):
                total += 1
                output += f"{' ' if p==len(self.ggmk['subkeys']) else '|'}{' '*9}|\n"
                output += f"{' ' if p==len(self.ggmk['subkeys']) else '|'}{' '*9}{' ' if q==len(gmk['subkeys'])+1 else '|'}{'-'*9} MK-{p:02d}-{q:02d} <> Codigo:{mk['codigo']}\n"
                output += f"{' ' if p==len(self.ggmk['subkeys']) else '|'}{' '*9}{' ' if q==len(gmk['subkeys']) else '|'}{' '*10}|\n"

                if not detalle:
                    output += f"{' ' if p==len(self.ggmk['subkeys']) else '|'}{' '*9}{' ' if q==len(gmk['subkeys']) else '|'}{' '*10}Unicas: {len(mk['subkeys'])}\n"
                    totalx += len(mk["subkeys"])
                else:
                    for r, key in enumerate(mk["subkeys"], start=1):
                        totalx += 1
                        output += f"{' ' if p==len(self.ggmk['subkeys']) else '|'}{' '*9}{' ' if q==len(gmk['subkeys']) else '|'}{' '*10}{' ' if r==len(mk['subkeys'])+1 else '|'}{'-'*9}{key['secuencia']} <> Codigo:{key['codigo']} - Cilindro:{key['cilindro']:<30} ({cantidad_master_pines(key['cilindro'])} MP) | Validaciones: self.ggmk:{valida_llave_abre_cilindro(self.ggmk['codigo'], key['cilindro'])} GMK:{valida_llave_abre_cilindro(gmk['codigo'], key['cilindro'])} MK:{valida_llave_abre_cilindro(mk['codigo'], key['cilindro'])} UNICA:{self.valida_llave_es_unica(mk['codigo'])} CRUZADA:{self.valida_llave_no_cruzada(key['cilindro'],[self.ggmk['codigo'],gmk['codigo'],mk['codigo']])}\n"

        output += f"\nTotal Llaves: {total+totalx:,}\n"
        output += f"Total Puertas: {totalx:,}\n"

        return output

    def carga_libro(self, ggmk):

        self.ggmk = ggmk

        # crear variables todas maestras y llaves usadas de la informacion del libro
        self.llaves_usadas = []
        self.todas_maestras = [self.ggmk["codigo"]]
        for gmk in self.ggmk["subkeys"]:
            self.todas_maestras.append(gmk["codigo"])
            for mk in gmk["subkeys"]:
                self.todas_maestras.append(mk["codigo"])
                for k in mk["subkeys"]:
                    self.llaves_usadas.append(k["codigo"])

        # eliminar botones de Rehacer y Guardar del menu
        self.buttons[1].destroy()
        self.buttons[3].destroy()

        arbol = self.genera_texto_arbol(detalle=False)
        self.muestra_arbol(arbol)

    def muestra_arbol(self, arbol):
        self.text_area.delete("1.0", "end")
        self.text_area.insert(END, arbol)

    def valida_llave_es_unica(self, llave):
        if self.llaves_usadas.count(llave) == 1:
            return 1
        return 0

    def valida_llave_no_cruzada(self, cilindro, mis_maestras):
        for maestra in self.todas_maestras:
            if maestra in mis_maestras:
                continue
            if valida_llave_abre_cilindro(maestra, cilindro):
                return 0

        return 1
